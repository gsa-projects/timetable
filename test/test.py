from src.timetable import *

from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

students = StudentList.load(
    '../data/2학년 학생별 시간표.xlsx',
    '../data/과목별 다교사수업.xlsx')

records: defaultdict[Subject, defaultdict[str, StudentList]] \
    = defaultdict(lambda: defaultdict(StudentList))

for student in students:
    for subject in student.classes:
        if student not in records[subject.name][subject.nth]:
            records[subject.name][subject.nth].append(student)

workbook = Workbook()
worksheet = workbook.active

def classes(sheet):
    row = 1
    max_row = -1

    for subject, dictionary in records.items():
        max_row += 2
        sheet.cell(row=max_row, column=1, value=str(subject))
        sheet.cell(row=max_row, column=1).font = Font(bold=True)
        row = max_row + 1

        for nth, student_list in dictionary.items():
            sheet.cell(row=row, column=nth, value=f'{nth}반')

            for i, student in enumerate(student_list):
                sheet.cell(row=row + i + 1, column=nth, value=student.name)
                max_row = max(max_row, row + i + 1)

def overlap_each_other(sheet):
    overlaps = defaultdict(set)

    for student1 in students:
        for student2 in students:
            if student1 == student2:
                continue

            overlap = student1.classes & student2.classes

            if sum(overlap.to_time()) >= 20:
                overlaps[sum(overlap.to_time())].add(frozenset([student1, student2]))

    r = 1
    for time in sorted(overlaps.keys(), reverse=True):
        student_list = overlaps[time]
        sheet.cell(row=r, column=1, value=f'{time}시수')
        sheet.cell(row=r, column=1).font = Font(bold=True)

        for i, (student1, student2) in enumerate(student_list):
            sheet.cell(row=r, column=i + 2, value=f"{student1.name} & {student2.name}")
            sheet.column_dimensions[get_column_letter(i + 2)].width = 15
        r += 1

def overlap_each(sheet):
    for i, student in enumerate(students, start=1):
        ranking = []

        for other in students:
            if student == other:
                continue

            overlap = student.classes & other.classes
            ranking.append((other, sum(overlap.to_time())))

        ranking.sort(key=lambda x: x[1], reverse=True)
        ranking = ranking[:5]

        sheet.cell(row=i, column=1, value=student.name)
        sheet.cell(row=i, column=1).font = Font(bold=True)

        for j, (other, time) in enumerate(ranking):
            sheet.cell(row=i, column=j + 2, value=f"{other.name} ({time}시수)")
            sheet.column_dimensions[get_column_letter(j + 2)].width = 15

classes(workbook.create_sheet('분반'))
overlap_each_other(workbook.create_sheet('서로 중복'))
overlap_each(workbook.create_sheet('개별 중복'))

del workbook['Sheet']
workbook.save('../data/시간표 분석.xlsx')
print('시간표 분석이 완료되었습니다.')
