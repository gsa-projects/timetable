import tkinter as tk
import pyautogui
import sv_ttk
from tkinter import ttk, filedialog
from PIL import Image
from openpyxl import load_workbook

from timetable import *

class Button(tk.Button):
    def __init__(self, master, memo, **kwargs):
        tk.Button.__init__(self, master, **kwargs)

        self.memo = memo    # 따로 인스턴스들을 저장할 딕셔너리

class ClassBlock(ttk.Frame):
    def __init__(self, parent, height=None, width=None, text="", command=None, style='TButton',
                 bg=('#eeeeee', '#cccccc'), fg='#000000', **kwargs):
        ttk.Frame.__init__(self, parent, height=height, width=width, style=style)

        self.pack_propagate(False)
        self.btn = Button(
            self,
            memo=kwargs,
            bg=bg[0],
            fg=fg,
            relief='flat',
            text=text,
            command=command,
            font=("Pretendard Variable", 11),
            borderwidth=0,
            activebackground=bg[1],
            activeforeground=fg,
            justify='left'      # 텍스트 왼쪽 정렬
        )
        self.btn.pack(fill=tk.BOTH, expand=1)

def capture(root: tk.Toplevel):
    """
    `root`의 스크린샷을 찍어 저장합니다. Windows 11에서는 탭이 둥근 사각형이기 때문에 이를 보정하기 위해
    `left_mask`와 `right_mask`를 사용해 좌우 하단을 보정합니다.
    """

    x, y, w, h = root.winfo_x(), root.winfo_y(), root.winfo_width(), root.winfo_height()

    # 윈도우 타이틀바 높이와 캡처 오차 수동 보정
    y += 57
    x += 3

    grabbed = pyautogui.screenshot(region=(x, y, w, h))

    left_mask = Image.new('RGB', (12, 12), '#eeeeee')
    right_mask = Image.new('RGB', (12, 12), '#fafafa')

    grabbed.paste(left_mask, (0, h - left_mask.height))
    grabbed.paste(right_mask, (w - right_mask.width, h - right_mask.height))
    grabbed.save(f'../images/{root.title()}.png')

def split(s, chuck_size=5):
    """
    문자열 `s`를 `chunk_size` 길이로 나눕니다. 다만, 영어로 된 문자열은 `chunk_size` 길이에서 고려되지 않고 더해집니다.
    '일반물리학II' 같은 문자열이 `chunk_size`가 5라면 '일반물리학\nII'로 나누어지는 것을 방지하기 위함입니다.
    """

    result = ""
    count = 0

    for char in s:
        if re.match(r'[ㄱ-ㅎㅏ-ㅣ가-힣]', char):
            count += 1
            if count == chuck_size + 1:
                result += '\n'
                count = 0

        result += char

    return result.strip('\n')

class App(ttk.Frame):
    def __init__(self, parent: tk.Tk):
        super().__init__()

        self.parent = parent
        self.parent_width = 300
        self.parent.geometry(f"{self.parent_width}x170")

        self.period_width = 45
        self.default_width = 125

        self.week_height = 50
        self.default_height = 70

        self.students = None

        self.title_screen()

    def title_screen(self):
        self.parent.title("Search")

        # 로고
        logo = ttk.Label(self, text='Timetable', font=("Product Sans Bold", 25))
        logo.place(x=self.parent_width // 2, y=50, anchor='center')

        # 경고 메시지
        warn = ttk.Label(self, font=("Pretendard Variable", 9))
        warn.place(x=self.parent_width // 2, y=145, anchor='center')

        # 파일 불러오기 버튼
        button = ttk.Button(self, text='Load Files', command=lambda: self.file_loading_screen(button, warn),
                            style='Accent.TButton')
        button.place(x=self.parent_width // 2, y=115, anchor='center')

    def file_loading_screen(self, button: ttk.Button, warn: ttk.Label):
        # 파일 선택, 2개의 .xlsx 파일
        f = filedialog.askopenfilenames(filetypes=[("강의실 엑셀 파일 & 전체 학생 시간표 파일", '*.xlsx')])

        if len(f) != 2:
            return
        else:
            button.configure(text='Loading...', state='disabled')

        # 2개의 파일에서 자동으로 timetable 파일과 subject 파일을 구분
        wb = load_workbook(f[0], read_only=True)
        if '2학기 강의실' in wb.sheetnames:
            subject_path = f[0]
            timetable_path = f[1]
        else:
            subject_path = f[1]
            timetable_path = f[0]

        # 학생 목록을 불러옴
        self.students = StudentList.load(timetable_path, subject_path)
        button.destroy()

        # 검색용 엔트리 생성
        name_var = tk.StringVar(value='학번이나 이름을 입력하세요')  # 엔트리 플레이스홀더
        name_entry = ttk.Entry(self, textvariable=name_var)
        name_entry.bind('<Return>', lambda _: self.search_screen(name_var, warn))
        name_entry.bind('<Button-1>', lambda _: name_var.set(''))
        name_entry.place(x=self.parent_width // 2, y=115, anchor='center')

    def search_screen(self, entry: tk.StringVar, warn: ttk.Label):
        name = entry.get().strip()
        if self.students is None:
            warn.configure(text='학생 목록을 불러와주세요')
            return

        student = self.students[name]
        if self.students is not None and student is not None:
            warn.configure(text='')

            new = tk.Toplevel(self.parent)
            new.title(f"{student}의 시간표")
            new.geometry(
                f"{self.period_width + len(Week) * self.default_width}x{self.week_height + (max_period - 1) * self.default_height}")
            new.resizable(False, False)

            menubar = tk.Menu(new)
            menubar.add_command(label="Screenshot", command=lambda:  capture(new))
            new.configure(menu=menubar)

            self.timetable_screen(new, student)
        else:
            warn.configure(text='학생을 찾을 수 없습니다')

    def timetable_screen(self, new: tk.Toplevel, student: Student):
        # 요일 행과 교시 열 사이 빈틈 마스크 생성
        block = ClassBlock(new, width=self.period_width, height=self.week_height)
        block.place(x=0, y=0)

        # 교시 열 생성
        for i in range(1, max_period):
            period_block = ClassBlock(new, text=f'{i}', width=self.period_width, height=self.default_height)
            period_block.place(x=0, y=self.week_height + (i - 1) * self.default_height)

        # 요일 행 생성
        for week in Week:
            week_block = ClassBlock(new, text=week.name, width=self.default_width, height=self.week_height)
            week_block.place(x=self.period_width + week.value * self.default_width, y=0)

        # 시간표 생성
        for week in Week:
            items = student.timetable[week].items()
            start: tuple[int, Subject] = None

            for i, ((_, period), subject) in enumerate(items):
                if start is None:
                    start = period, subject

                # 연강 여부 확인
                if subject != start[1]:
                    duration = period - start[0]  # n연강
                else:
                    continue

                if start[1] == GAP:
                    start = period, subject
                    continue

                # class block 생성
                subject_name = start[1].name.replace(' ', '')
                label_txt = split(subject_name, 5)
                class_block = ClassBlock(new, width=self.default_width, height=self.default_height * duration,
                                         text=label_txt, bg=start[1].type.color,
                                         label_txt=label_txt, subject=start[1])
                class_block.place(x=self.period_width + week.value * self.default_width,
                                  y=self.week_height + (start[0] - 1) * self.default_height)

                # class block을 누르면 강의 정보(선생님, 강의실, 분반)를 보여줌
                def on_class_block_press(event):
                    source: Button = event.widget
                    got: Subject = source.memo['subject']

                    source.configure(
                        text=f'{got.teacher.name}T\n{got.teacher.classroom}\n{got.nth}분반')
                class_block.btn.bind('<ButtonPress>', on_class_block_press)

                # class block을 놓으면 다시 강의 이름을 보여줌
                def on_class_block_release(event):
                    source: Button = event.widget
                    got = source.memo['label_txt']

                    source.configure(text=got)
                class_block.btn.bind('<ButtonRelease>', on_class_block_release)

                start = period, subject

if __name__ == "__main__":
    root = tk.Tk()

    # 테마 및 스타일 설정
    sv_ttk.set_theme("light")
    style = ttk.Style()
    style.configure('TButton', font=("Pretendard Variable", 12))

    # 앱 생성
    app = App(root)
    app.pack(fill="both", expand=True)

    # 창 중앙에 배치
    root.update()
    root.minsize(root.winfo_width(), root.winfo_height())
    x_coordinate = int((root.winfo_screenwidth() / 2) - (root.winfo_width() / 2))
    y_coordinate = int((root.winfo_screenheight() / 2) - (root.winfo_height() / 2))
    root.geometry("+{}+{}".format(x_coordinate, y_coordinate - 20))

    root.mainloop()
